import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\Admin\Downloads\Trả lời sự kiện (Câu trả lời).xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n=== Sheet: {name} ({ws.max_row} rows x {ws.max_column} cols) ===')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            print(f'Row {i+1}:', row)